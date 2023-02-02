import openpyxl
import csv
import datetime

csvfile = open('output.csv', mode='w', newline='')
csvwriter = csv.writer(csvfile)
csvwriter.writerow(['section', 'table name', 'category', 'value', 'year', 'timestamp'])

workbook = openpyxl.load_workbook('GdpByInd/ValueAdded.xlsx')
sheets = workbook.sheetnames
# print(sheets)

# Date
today_date = datetime.date.today()

# Select sheet 1
sheet = workbook[sheets[1]]

# print(sheet.max_row, sheet.max_column)
cell = sheet.cell(row=1, column=1)
table = cell.value.strip()

# Select a cell
for row in range(9, 110):
    cell = sheet.cell(row=row, column=2)
    category = cell.value.strip()

    for col in range(4, 28):
        cell = sheet.cell(row=row, column=col)
        value = cell.value

        cell = sheet.cell(row=8, column=col)
        year = cell.value

        csvwriter.writerow(['ValueAdded', table, category, value, year, today_date])

csvfile.close()
